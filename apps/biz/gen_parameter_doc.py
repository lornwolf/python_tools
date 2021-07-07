from copy import copy

import cx_Oracle
import openpyxl


def main():
    """
    从DB获取2.0.1版的所有参数，跟2.2.1设计书中的参数作比较，如果有新增的参数，就出力到文件中。
    """

    # 模块名字典。
    modules = {"会計コア":"FC", "一般会計":"GL", "債権債務":"PR"}
    # 2.0.1版所有的参数（从DB获取）。
    parameters_201 = []
    # 2.2.1版新增的参数。
    parameters_221 = []

    # 连接Oracle，获取2.0.1版所有的参数名称。
    db = cx_Oracle.connect('SIBH01/Sibhpass15@192.168.243.214:1521/SAZABY')
    cursor = db.cursor()
    cursor.execute('SELECT PARAMETER_CD FROM FMS_FC_PARAMETER')
    all_data = cursor.fetchall()
    print("既存パラメータ全部 " + str(len(all_data)) + " 件。")
    for index in range(len(all_data)):
        parameters_201.append(all_data[index][0])
        # print(all_data[index][0])
    db.close()

    # 忽略打开Excel时的警告。
    openpyxl.reader.excel.warnings.simplefilter('ignore')

    # 打开Excel模板文件。
    wb_for_output = openpyxl.load_workbook('/01_input/移行シート_パラメータ管理.xlsx')
    ws_for_output = wb_for_output.worksheets[0]
    # 模板文件的开始写入位置。
    row_start = 12

    # 打开参数设计书文件。
    workbook = openpyxl.load_workbook('/01_input/05_パラメータ一覧.xlsx', read_only=True, keep_vba=False)
    # 遍历所有Sheet。
    for sheet in workbook:
        # 设计书文件的开始读取位置。
        row_no = 4
        # 从指定位置遍历所有行。
        for row in sheet.iter_rows(min_row=5):
            row_no += 1
            # 判断最终版本的记述，如有的话该行数据跳过。
            last_version = sheet.cell(row=row_no, column=22).value
            if last_version is not None and last_version.startswith("Ver"):
                continue
            # 获取参数名。
            parameter_name =sheet.cell(row=row_no, column=3).value
            # 忽略空值。
            if parameter_name is None:
                continue

            # 判断该参数在2.0.1版本中存不存在。
            if parameter_name not in parameters_201 and parameter_name not in parameters_221:
                parameters_221.append(parameter_name)

                # 判断是不是仕訳パターンコード。
                pattern_code = sheet.cell(row=row_no, column=18).value
                if pattern_code is not None and pattern_code == "仕訳パターンコード":
                    print("仕訳パターンコード：" + parameter_name)

                # パラメータコード
                ws_for_output.cell(row=row_start, column=2).value = parameter_name
                # モジュール分類
                ws_for_output.cell(row=row_start, column=4).value = modules[sheet.cell(row=row_no, column=2).value]
                # モジュール名
                ws_for_output.cell(row=row_start, column=5).value = sheet.cell(row=row_no, column=2).value
                # パラメータ名
                ws_for_output.cell(row=row_start, column=7).value = sheet.cell(row=row_no, column=6).value
                # パラメータ説明
                ws_for_output.cell(row=row_start, column=8).value = sheet.cell(row=row_no, column=8).value
                # システムデフォルト値
                ws_for_output.cell(row=row_start, column=9).value = sheet.cell(row=row_no, column=9).value
                # 从第二行开始，复制上一行的单元格样式。（不能复制行，真麻烦）
                if row_start > 12:
                    for i in range(1, 13):
                        new_cell = ws_for_output.cell(row=row_start, column=i)
                        old_cell = ws_for_output.cell(row=row_start - 1, column=i)
                        if old_cell.has_style:
                            new_cell._style =  copy(old_cell._style)
                row_start += 1

    workbook.close()
    wb_for_output.save("C:/02_output/result.xlsx")
    wb_for_output.close()

    # 打印所有新增参数。
    for index in range(len(parameters_221)):
        print(parameters_221[index])


if __name__ == '__main__':
    main()